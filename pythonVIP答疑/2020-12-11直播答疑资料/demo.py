import  openpyxl

openpyxl.Workbook()  # Workbook()是openpyxl中的一个类

#print(dir(openpyxl.Workbook()))
print(len(dir(openpyxl.Workbook())))  # len是Python中的内置函数，用于查看列表或字符串中字符的长度
print(len('hello'))  # 5  len就是length的缩写

import docx    # Python中有十几万个第三方模块，安装第三方模块，要求每个人必须会
docx.Document()  # Document()是docx中的一个函数

class Student(object):
    school='北京马士兵教育'  #类属性   ,为什么不能写成self.school呢？因为类对象产生时，还没有实例对象self

    # def __init__(self):
    #     self.name='张三'  #属性  这个Student类的所有对象的名字都是张三

    def __init__(self,name):  #name是__init__这个方法的参数  等同于方法的局部变量
        self.name=name   # =左侧self.name是类的实例对象的属性 , 右侧   name 局部变量
                        # 将局部变量name的值赋值给了类的实例对象的属性  。 在创建完类的实例对象之后赋的值


    def fun(self): #方法

        pass

print(type(openpyxl))  # 查看openpyxl的类型
print(type(docx)) #<class 'module'>
print(id(openpyxl)) #2093473982160  openpyxl在内存中的内存地址
print(type(Student)) #<class 'type'>
print(id(Student)) #1879882516208
print('---------------------------------')
for i in range(10):
    stu=Student() #每执行一次，创建一个学生对象
    print(stu.name)
    stu.name='李四--'+str(i)
    print(stu.name)









