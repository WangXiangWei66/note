# 教育机构 ：马士兵教育
# 讲    师：杨淑娟
#使用Python操作Excel，使用到openpyxl模块
import  openpyxl
#加载Excel文件
wk=openpyxl.load_workbook('update.xlsx')  #wk 工作薄对象
sheet=wk.active  #获取活动sheet页   工作表对象
#将Excel中的数据读取到Python程序中来
#print(sheet.max_row)    #一共有11行

# 要修改的数据为
product_price={'Garlic':3.07,'Watermelon':1.17,'Celery':1.19}

for row_num in range(1,sheet.max_row):
    produce_name=sheet.cell(row=row_num,column=1).value
    #print(produce_name)
    if produce_name in product_price:  #如果produce_name在字典中存储，说明它就是要修改的数据
        sheet.cell(row=row_num,column=2).value=product_price[produce_name]

#保存
wk.save('update.xlsx')



