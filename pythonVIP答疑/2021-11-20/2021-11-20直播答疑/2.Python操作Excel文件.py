# coding:utf-8
# 教育机构：马士兵教育
# 讲    师：杨淑娟
import openpyxl
lst=[
    ['姓名','语文','数学'],
    ['张三',90,89],
    ['李四',89,78]
]

lst2=[
    ['姓名1','语文1','数学1'],
    ['张三1',90,89],
    ['李四1',89,78]
]

#将两个列表中的数据存储到同一个Excel文件里的不同sheet页（工作表）

wk=openpyxl.Workbook() #创建一个新的工作簿(一个Excel文件）
#创建２张工作表
sheet1=wk.create_sheet('Sheet1')
sheet2=wk.create_sheet('Sheet2')
#分别向两个工作表中写数据
for item in lst:
    sheet1.append(item)

for item in lst2:
    sheet2.append(item)

# 保存工作簿
wk.save('my.xlsx')

