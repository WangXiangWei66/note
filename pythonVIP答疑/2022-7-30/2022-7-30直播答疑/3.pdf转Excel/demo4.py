#教育机构 ：马士兵教育
#讲    师：杨淑娟
import os
import  pdfplumber    # 第三方库 操作pdf文件
import  openpyxl   # 第三方库，操作Excel文件
file_names=os.listdir('./score')  # 获取score这个目录下所有的文件名称

lst_names=[]
for item in file_names:    #  过滤的操作  （把.pdf的文件提取出来）
    if item.endswith('.pdf'):   # 文件名中以.pdf结尾的即为PDF文件
        lst_names.append('score/'+item)
#print(lst_names)
lst=[]

#提取表格中的数据
for item in lst_names:
    with pdfplumber.open(item) as pdf:
        #print(len(pdf.pages))
        for p in range(0,len(pdf.pages)): # 从第一页开始打开，第一页的索引为0
            pdf_page=pdf.pages[p]    #获取单独的一个 pdf页
            #提取表格数据
            table=pdf_page.extract_table()  # 提取表格
            for i in table:
                if i[0]!=None and  i[0]!='学号' and  i[0]!='总\n评\n成\n绩\n分\n析':
                    lst.append(i)


score_lst=[]#用于存储及格学生信息
score_lst2=[]#用于存储不及格学生的信息
for item in lst:
    if item[4]>='60':
        score_lst.append(item[1:])  # 大于60 （及格的学生）score_lst
    else:
        score_lst2.append(item[1:])  #  不及格的学生

'''for item in score_lst2:
    print(item)
'''
for index,item in zip(range(1,len(score_lst)+1),score_lst):
    item.insert(0,index)

for index,item in zip(range(1,len(score_lst2)+1),score_lst2):
    item.insert(0,index)

'''for item in score_lst2:
    print(item)'''
# 在索引为0的位置 上插入 标题
score_lst.insert(0,['序号','姓名','平时成绩','考试成绩','总成绩','学分绩点','备注'])
score_lst2.insert(0,['序号','姓名','平时成绩','考试成绩','总成绩','学分绩点','备注'])

#保存到Excel文件中
workbook=openpyxl.Workbook()
sheetA=workbook.create_sheet()   #
sheetA.title='02微机原理及格学员名单'

sheetB=workbook.create_sheet()
sheetB.title='02微机原理不及格学员名单'


#分别向两个sheet中添加数据
for item in score_lst:
    sheetA.append(item)

for item in score_lst2:
    sheetB.append(item)

#保存Excel文件
workbook.save('02微机原理学员成绩统计.xlsx')