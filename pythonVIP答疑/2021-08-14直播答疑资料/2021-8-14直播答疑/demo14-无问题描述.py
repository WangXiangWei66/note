# coding:utf-8
# 教育机构：马士兵教育
# 讲    师：杨淑娟

import openpyxl as excel
from docx import Document
import os

def read(file):
    ex=excel.load_workbook(file)
    sheet=ex.active
    row=sheet.max_row
    col=sheet.max_column
    for i in range(1,row+1):
        lst_row=[]
        for j in range(1,col+1):
            data=sheet.cell(row=i,column=j).value
            if data is not None:
                lst_row.append(data)
        lst.append(lst_row)
    print(lst)
#   加if判断： [['序号', '项目名称', '批准文号', '采购\n预算\n（万元）', '交易平台', '交易方式', '发布公告\n时间\n（月/日）', ' 评审时间 （月/日）', '发布\n媒介', '中标成交金额\n（万元）', '备注', '成交单位', '申请科室', '清单', '供货期', '质保期'], [1, '口腔科医疗设备采购项目', '微分', 73.5, '二哥', '公开招标', 20170321, '2021年3月22日', '第三方招标代理机构', 50.75, '牙科治疗仪为法国赛特力5P NEWTRON XS、超声骨切割系统为法国赛特力PIEZOTOME SOLO LED、普通牙科治疗台为台湾光宇Frontier(开拓者）、种植牙科治疗台为台湾光宇K-7推车式、高档口腔科综合治疗台为台湾光宇K-7，空气压缩机为上海硅莱GA-85', '合肥洁皓贸易有限公司', '无法个个', '牙科治疗仪1套、超声骨切割系统1套、普通牙科治疗台2套、种植牙科治疗台2套、高档口腔科综合治疗台1套和空气压缩机1套', '合同签订后国产设备30天，进口设备60天内完成安装', '5年'], [2, '纷纷', '任务范围', '仍无法', '韦富文', '威锋网', 20170321, '2021年4月13日', 'Greg', 500, '仍无法', '任务分为非', '范文芳', '分为非', 6, 2], [0, 1, 2, 561.3, 4, 5, 6, '2021年5月15日', 8, 365, 10, 11, 12, 13, 14, 15]]

lst_row = []
lst = []
read("工作表(1).xlsx")